"""
审计取证单生成工具 (CLI版)
用法: python 取证单生成CLI.py --excel <Excel文件> --template <Word模板> [--output <输出目录>]

依赖: openpyxl, python-docx, lxml
安装: pip install openpyxl python-docx lxml
"""

import sys
import os
import argparse

try:
    import openpyxl
    from docx import Document
    from docx.oxml.ns import qn
    from docx.oxml import OxmlElement
    from copy import deepcopy
except ImportError as e:
    missing = str(e)
    print(f'错误：缺少依赖库。请先安装：pip install openpyxl python-docx lxml', file=sys.stderr)
    print(f'详细信息：{missing}', file=sys.stderr)
    sys.exit(1)


def chinese_num(n):
    """生成中文编号 （一）（二）..."""
    bases = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十',
             '十一', '十二', '十三', '十四', '十五', '十六', '十七', '十八', '十九', '二十',
             '二十一', '二十二', '二十三', '二十四', '二十五', '二十六', '二十七', '二十八', '二十九', '三十',
             '三十一', '三十二', '三十三', '三十四', '三十五', '三十六', '三十七', '三十八', '三十九', '四十']
    if n < len(bases):
        return f'（{bases[n]}）'
    return f'（{n + 1}）'


def read_excel(excel_path):
    """读取Excel数据"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cols = list(row)
        if len(cols) < 7 or cols[1] is None:
            continue
        data.append({
            'title': str(cols[2] or '').strip(),
            'qualitative': str(cols[3] or '').strip(),
            'description': str(cols[4] or '').strip(),
            'basis': str(cols[5] or '').strip(),
            'dept': str(cols[6] or '').strip(),
        })
    return data


def group_by_title(data):
    """按'问题标题'分组相同项"""
    groups = []
    current = None
    for row in data:
        if current is not None and current['title'] != row['title']:
            groups.append(current)
            current = {'title': row['title'], 'rows': [row]}
        else:
            if current is None:
                current = {'title': row['title'], 'rows': [row]}
            else:
                current['rows'].append(row)
    if current:
        groups.append(current)
    return groups


def build_paras(group):
    """构建 #2 占位符列表，含去重和合并逻辑"""
    subs = []
    cur = None
    for row in group['rows']:
        q = row['qualitative']
        if cur is not None and cur['q'] != q:
            subs.append(cur)
            cur = {'q': q, 'rows': [row]}
        else:
            if cur is None:
                cur = {'q': q, 'rows': [row]}
            else:
                cur['rows'].append(row)
    if cur:
        subs.append(cur)

    result = []
    for i, sub in enumerate(subs):
        if sub['q']:
            title_text = f'{chinese_num(i)}{sub["q"]}'
        else:
            title_text = chinese_num(i)
        result.append(title_text)

        rows = sub['rows']
        violation_texts = []
        for row in rows:
            basis = row['basis']
            if basis:
                violation_texts.append(f'该事项违反了{basis}')
            else:
                violation_texts.append(None)

        deduped = list(violation_texts)
        j = 0
        while j < len(deduped):
            if deduped[j] is not None:
                k = j + 1
                while k < len(deduped) and deduped[k] == deduped[j]:
                    k += 1
                if k > j + 1:
                    for m in range(j, k - 1):
                        deduped[m] = None
                    deduped[k - 1] = deduped[k - 1].replace('该事项违反了', '以上事项违反了')
                    j = k
                else:
                    j += 1
            else:
                j += 1

        for idx, row in enumerate(rows):
            for line in row['description'].split('\n'):
                line = line.strip()
                if line:
                    result.append(line)
            if deduped[idx] is not None:
                result.append(deduped[idx])
            result.append(f'责任部门：{row["dept"]}')

    return result


def find_cells_in_tbl(tbl_elem):
    """在表格XML中查找 #1 和 #2 占位元格"""
    tc_1 = None
    tc_2 = None
    for tr in tbl_elem.findall(qn('w:tr')):
        for tc in tr.findall(qn('w:tc')):
            texts = []
            for t in tc.iter(qn('w:t')):
                if not t.text:
                    continue
                texts.append(t.text)
            full_text = ''.join(texts).strip()
            if full_text == '#1' and tc_1 is None:
                tc_1 = tc
            if '#2' in full_text and tc_2 is None:
                tc_2 = tc
    return tc_1, tc_2


def set_cell_text(tc_elem, text):
    """替换单元格为新文本，保留格式"""
    paras = tc_elem.findall(qn('w:p'))
    if not paras:
        return
    for p in paras[1:]:
        tc_elem.remove(p)

    p0 = paras[0]
    orig_runs = p0.findall(qn('w:r'))
    rPr_template = orig_runs[0].find(qn('w:rPr')) if orig_runs else None
    for r in p0.findall(qn('w:r')):
        p0.remove(r)

    r = OxmlElement('w:r')
    if rPr_template is not None:
        r.insert(0, deepcopy(rPr_template))
    t_elem = OxmlElement('w:t')
    t_elem.text = text
    t_elem.set(qn('xml:space'), 'preserve')
    r.append(t_elem)
    p0.append(r)


def fill_cell_multi(tc_elem, para_texts):
    """填多段落到单元格，保留格式"""
    paras = tc_elem.findall(qn('w:p'))
    if not paras:
        return

    ref_p = paras[0]
    orig_runs = ref_p.findall(qn('w:r'))
    rPr_template = orig_runs[0].find(qn('w:rPr')) if orig_runs else None
    for p in paras:
        tc_elem.remove(p)

    for text in para_texts:
        new_p = deepcopy(ref_p)
        for r in new_p.findall(qn('w:r')):
            new_p.remove(r)
        if text:
            r = OxmlElement('w:r')
            if rPr_template is not None:
                r.insert(0, deepcopy(rPr_template))
            t_elem = OxmlElement('w:t')
            t_elem.text = text
            t_elem.set(qn('xml:space'), 'preserve')
            r.append(t_elem)
            new_p.append(r)
        tc_elem.append(new_p)


def fill_table(tbl_elem, group):
    """填充一个表格"""
    tc_1, tc_2 = find_cells_in_tbl(tbl_elem)
    if tc_1 is None or tc_2 is None:
        return False
    set_cell_text(tc_1, group['title'])
    fill_cell_multi(tc_2, build_paras(group))
    return True


def generate(excel_path, template_path, output_dir):
    """生成合并取证单，返回 (输出路径, 分组数, 总行数)"""
    data = read_excel(excel_path)
    if not data:
        print('错误：Excel中没有找到有效数据行！请确保第二列有序号，第三列有问题标题。', file=sys.stderr)
        sys.exit(1)

    groups = group_by_title(data)
    doc = Document(template_path)
    body = doc.element.body

    template_body_children = [deepcopy(child) for child in body]

    first_tbl = doc.tables[0]
    fill_table(first_tbl._tbl, groups[0])

    for i in range(1, len(groups)):
        group = groups[i]
        new_elements = [deepcopy(child) for child in template_body_children]
        for elem in new_elements:
            if elem.tag == qn('w:tbl'):
                fill_table(elem, group)
                break

        p_break = OxmlElement('w:p')
        r_break = OxmlElement('w:r')
        br = OxmlElement('w:br')
        br.set(qn('w:type'), 'page')
        r_break.append(br)
        p_break.append(r_break)
        body.append(p_break)

        for elem in new_elements:
            body.append(elem)

    for tbl in body.findall(qn('w:tbl')):
        for tr in tbl.findall(qn('w:tr')):
            for tc in tr.findall(qn('w:tc')):
                paras = tc.findall(qn('w:p'))
                for p in paras:
                    has_text = any(t.text and t.text.strip() for t in p.iter(qn('w:t')))
                    if not has_text:
                        tc.remove(p)

    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, '审计取证单_合并.docx')
    doc.save(output_path)
    return output_path, len(groups), len(data)


def main():
    parser = argparse.ArgumentParser(
        description='审计取证单生成工具 - 将Excel问题清单按大类填入Word取证单模板，生成合并取证单'
    )
    parser.add_argument('--excel', '-e', required=True,
                        help='Excel 问题清单文件路径')
    parser.add_argument('--template', '-t', required=True,
                        help='Word 取证单模板路径（含 #1 和 #2 占位符）')
    parser.add_argument('--output', '-o', default=None,
                        help='输出目录（默认：Excel同目录下的"生成后大取证单"文件夹）')
    args = parser.parse_args()

    if not os.path.exists(args.excel):
        print(f'错误：Excel 文件不存在：{args.excel}', file=sys.stderr)
        sys.exit(1)
    if not os.path.exists(args.template):
        print(f'错误：模板文件不存在：{args.template}', file=sys.stderr)
        sys.exit(1)

    output_dir = args.output
    if not output_dir:
        excel_dir = os.path.dirname(os.path.abspath(args.excel))
        output_dir = os.path.join(excel_dir, '生成后大取证单')

    try:
        output_path, num_groups, num_rows = generate(args.excel, args.template, output_dir)
        print(f'[OK] 生成完成')
        print(f'     总记录数：{num_rows}')
        print(f'     取证单数：{num_groups}')
        print(f'     输出文件：{output_path}')
    except Exception as e:
        print(f'错误：{e}', file=sys.stderr)
        sys.exit(1)


if __name__ == '__main__':
    main()
